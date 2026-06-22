# EXCEL plugin, imports all the plug in data into excel to be used as a single research notebook
# Comoletely obvious and completly self-explanatory, no comments needed

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BLUE = Font(name='Arial', size=10, color='0000FF')
BLACK = Font(name='Arial', size=10, color='000000')
BLACK_BOLD = Font(name='Arial', size=10, color='000000', bold=True)
WHITE_BOLD = Font(name='Arial', size=11, color='FFFFFF', bold=True)
TITLE_FONT = Font(name='Arial', size=14, bold=True)
SUBTITLE_FONT = Font(name='Arial', size=9, italic=True, color='666666')
SECTION_FILL = PatternFill('solid', start_color='1F4E78', end_color='1F4E78')
TARGET_FILL = PatternFill('solid', start_color='FFF2CC', end_color='FFF2CC')
SUBTOTAL_TOP_BORDER = Border(top=Side(style='thin'))
PCT_FMT = '0.0%;(0.0%);"-"'
NUM_FMT = '$#,##0;($#,##0);"-"'


def _section_header(ws, row, title, n_cols, start_col=1):
    cell = ws.cell(row=row, column=start_col, value=title)
    cell.font = WHITE_BOLD
    cell.fill = SECTION_FILL
    for c in range(start_col + 1, start_col + n_cols):
        ws.cell(row=row, column=c).fill = SECTION_FILL
    return row + 1


def build_competitive_landscape_sheet(wb, ticker, names, timeline, revenue_by_ticker,
                                       share_by_ticker, unmatched_private_competitors,
                                       sic_description=None):
    ws = wb.create_sheet("Competitive Landscape")
    n_years = len(timeline)
    first_col = 2

    ws.cell(row=1, column=1, value=f"Competitive Landscape — {names.get(ticker, ticker)} ({ticker})").font = TITLE_FONT
    subtitle = "Share = company revenue / combined revenue of identified peer group (not total industry size)"
    if sic_description:
        subtitle += f"  |  Industry: {sic_description}"
    ws.cell(row=2, column=1, value=subtitle).font = SUBTITLE_FONT

    row = 4
    ws.cell(row=row, column=1, value="Company").font = BLACK_BOLD
    for i, yr in enumerate(timeline):
        c = ws.cell(row=row, column=first_col + i, value=yr)
        c.font = BLACK_BOLD
        c.alignment = Alignment(horizontal='center')
    ws.cell(row=row, column=first_col + n_years, value="Trend").font = BLACK_BOLD
    row += 1

    row = _section_header(ws, row, "REVENUE ($MM)", n_years + 1)
    tickers_sorted = sorted(revenue_by_ticker.keys(), key=lambda t: revenue_by_ticker[t].get(timeline[-1], 0), reverse=True)
    for tk in tickers_sorted:
        is_target = (tk == ticker)
        label = f"{names.get(tk, tk)} ({tk})" + ("  ← target" if is_target else "")
        lbl_cell = ws.cell(row=row, column=1, value=label)
        lbl_cell.font = BLACK_BOLD if is_target else BLACK
        for i, yr in enumerate(timeline):
            val = revenue_by_ticker[tk].get(yr)
            cell = ws.cell(row=row, column=first_col + i)
            if val is not None:
                cell.value = val
            cell.font = BLUE
            cell.number_format = NUM_FMT
            if is_target:
                cell.fill = TARGET_FILL
        if is_target:
            lbl_cell.fill = TARGET_FILL
        row += 1

    row += 1
    row = _section_header(ws, row, "SHARE OF PEER-GROUP REVENUE", n_years + 1)
    ticker_row_map = {}
    for tk in tickers_sorted:
        is_target = (tk == ticker)
        label = f"{names.get(tk, tk)} ({tk})" + ("  ← target" if is_target else "")
        lbl_cell = ws.cell(row=row, column=1, value=label)
        lbl_cell.font = BLACK_BOLD if is_target else BLACK
        first_share_col = None
        last_share_col = None
        for i, yr in enumerate(timeline):
            val = share_by_ticker.get(tk, {}).get(yr)
            cell = ws.cell(row=row, column=first_col + i)
            if val is not None:
                cell.value = val
                if first_share_col is None:
                    first_share_col = first_col + i
                last_share_col = first_col + i
            cell.font = BLACK_BOLD if is_target else BLACK
            cell.number_format = PCT_FMT
            if is_target:
                cell.fill = TARGET_FILL
        trend_cell = ws.cell(row=row, column=first_col + n_years)
        if first_share_col is not None and last_share_col is not None and first_share_col != last_share_col:
            trend_cell.value = f"={get_column_letter(last_share_col)}{row}-{get_column_letter(first_share_col)}{row}"
        trend_cell.font = BLACK_BOLD if is_target else BLACK
        trend_cell.number_format = '+0.0%;-0.0%;"-"'
        if is_target:
            lbl_cell.fill = TARGET_FILL
            trend_cell.fill = TARGET_FILL
        ticker_row_map[tk] = row
        row += 1

    if unmatched_private_competitors:
        row += 1
        ws.cell(row=row, column=1, value="Named as competitors but not found in SEC registry (likely private — no comparable financials):").font = SUBTITLE_FONT
        row += 1
        for name in unmatched_private_competitors:
            ws.cell(row=row, column=1, value=f"  • {name}").font = BLACK
            row += 1

    ws.column_dimensions['A'].width = 38
    for i in range(n_years + 1):
        ws.column_dimensions[get_column_letter(first_col + i)].width = 12
    return ws, ticker_row_map


def build_industry_factors_sheet(wb, ticker, company_name, filing_date,
                                  product_type_result, end_markets, new_risk_sentences):
    ws = wb.create_sheet("Industry Factors")
    ws.cell(row=1, column=1, value=f"Industry Factors — {company_name} ({ticker})").font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"Derived from 10-K filed {filing_date} (Item 1 Business + Item 1A Risk Factors)").font = SUBTITLE_FONT

    row = 4
    row = _section_header(ws, row, "PRODUCT TYPE READ (HEURISTIC)", 3)
    ws.cell(row=row, column=1, value="Classification").font = BLACK_BOLD
    ws.cell(row=row, column=2, value=product_type_result['classification']).font = BLACK
    row += 1
    ws.cell(row=row, column=1, value="Consumable/recurring language hits").font = BLACK
    ws.cell(row=row, column=2, value=product_type_result['consumable_score']).font = BLUE
    row += 1
    ws.cell(row=row, column=1, value="Durable/one-time language hits").font = BLACK
    ws.cell(row=row, column=2, value=product_type_result['durable_score']).font = BLUE
    row += 1
    ws.cell(row=row, column=1, value="Note: keyword heuristic on filing text — verify against your own read of the business").font = SUBTITLE_FONT
    row += 2

    row = _section_header(ws, row, "END MARKETS MENTIONED IN FILING", 3)
    if end_markets:
        for m in end_markets:
            ws.cell(row=row, column=1, value=f"  • {m.title()}").font = BLACK
            row += 1
    else:
        ws.cell(row=row, column=1, value="  (none of the common end-markets list matched — may need manual review)").font = BLACK
        row += 1
    row += 1

    row = _section_header(ws, row, "NEW RISK LANGUAGE VS. PRIOR YEAR'S 10-K", 3)
    ws.cell(row=row, column=1,
            value="Sentences with no close match in the prior year's Risk Factors section — i.e. likely new/emerging concerns:").font = SUBTITLE_FONT
    row += 1
    if new_risk_sentences:
        for s in new_risk_sentences:
            cell = ws.cell(row=row, column=1, value=f"  • {s}")
            cell.font = BLACK
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            ws.row_dimensions[row].height = 30
            row += 1
    else:
        ws.cell(row=row, column=1, value="  (no prior-year filing available, or no meaningfully new language detected)").font = BLACK
        row += 1

    ws.column_dimensions['A'].width = 90
    ws.column_dimensions['B'].width = 16
    return ws


def build_scorecard_sheet(wb, ticker, company_name, share_trend, revenue_growth,
                           new_risk_count, product_type_result):
    ws = wb.create_sheet("Research Signal Summary")
    ws.cell(row=1, column=1, value=f"Research Signal Summary — {company_name} ({ticker})").font = TITLE_FONT
    ws.cell(row=2, column=1,
            value="Informational synthesis of the signals below — not a recommendation. Not financial advice.").font = SUBTITLE_FONT

    row = 4
    row = _section_header(ws, row, "SCORED SIGNALS", 3)
    ws.cell(row=row, column=1, value="Signal").font = BLACK_BOLD
    ws.cell(row=row, column=2, value="Reading").font = BLACK_BOLD
    ws.cell(row=row, column=3, value="Lean").font = BLACK_BOLD
    row += 1

    def lean_label(score):
        return {1: "+ favorable", 0: "neutral", -1: "− unfavorable"}.get(score, "neutral")

    share_score = 1 if share_trend > 0.05 else (-1 if share_trend < -0.05 else 0)
    growth_score = 1 if revenue_growth > 0.05 else (-1 if revenue_growth < 0 else 0)

    rows_data = [
        ("Peer-group share trend (last yr vs. first yr)", f"{share_trend:+.1%}", lean_label(share_score)),
        ("Latest YoY revenue growth", f"{revenue_growth:+.1%}", lean_label(growth_score)),
    ]
    for label, reading, lean in rows_data:
        ws.cell(row=row, column=1, value=label).font = BLACK
        ws.cell(row=row, column=2, value=reading).font = BLACK
        ws.cell(row=row, column=3, value=lean).font = BLACK
        row += 1

    total_score = share_score + growth_score
    row += 1
    overall = "Bullish lean" if total_score >= 1 else ("Bearish lean" if total_score <= -1 else "Neutral")
    cell = ws.cell(row=row, column=1, value=f"Overall lean: {overall}")
    cell.font = BLACK_BOLD
    cell.border = SUBTOTAL_TOP_BORDER
    row += 2

    row = _section_header(ws, row, "CONTEXT (NOT SCORED)", 3)
    ws.cell(row=row, column=1, value="Product type").font = BLACK
    ws.cell(row=row, column=2, value=product_type_result['classification']).font = BLACK
    row += 1
    ws.cell(row=row, column=1, value="New risk language this filing cycle").font = BLACK
    ws.cell(row=row, column=2, value=f"{new_risk_count} new statement(s) — see Industry Factors tab").font = BLACK
    row += 2

    ws.cell(row=row, column=1,
            value="This is a synthesis of structured SEC filing data and simple heuristics, not investment advice. "
                  "Verify drivers on the other tabs before acting on them.").font = SUBTITLE_FONT

    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 16
    return ws
