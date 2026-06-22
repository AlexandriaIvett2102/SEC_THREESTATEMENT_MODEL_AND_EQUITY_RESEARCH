# Import necessary libraries

import os
import sys
import json
import time
import platform
import subprocess
import datetime
import requests
import ipywidgets as widgets
from IPython.display import display, clear_output, FileLink

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =====================================================================
# SEC COMPLIANCE / REQUEST LAYER
# =====================================================================
# SEC requires a descriptive User-Agent with a real contact email, and asks
# that automated tools stay under ~10 requests/second. We also retry with
# backoff on 403/429 instead of silently treating a block page as "no data".
# Uses contact email to make requests to SEC URL
HEADERS = {'User-Agent': 'Sasha Brown s.i.2002@gmail.com'}
SESSION = requests.Session()
SESSION.headers.update(HEADERS)
# Delays requests to avoid SEC rate limiting (10/sec) and 429/403 blocks. Also
# adds a small backoff on retries to avoid hammering the SEC if they are
# already rate-limiting us.
REQUEST_DELAY = 0.15
TICKER_CACHE_PATH = 'sec_tickers_cache.json'
TICKER_CACHE_MAX_AGE = 86400  # 1 day

# Creates function to get data from SEC URL with retries and backoff, uses timeout
# and makes decisions based on SEC code response (403, 429, 200, etc.)
def sec_get(url, max_retries=4):
    last_resp = None
    for attempt in range(max_retries):
        resp = SESSION.get(url, timeout=20)
        last_resp = resp
        if resp.status_code == 200:
            time.sleep(REQUEST_DELAY)
            return resp
        if resp.status_code in (403, 429):
            time.sleep(REQUEST_DELAY * (2 ** attempt) + 1)
            continue
        return resp
    return last_resp

# Creates function to load ticker map from SEC, caches it locally for a day to avoid repeated requests
# If the cache is older than a day, it will fetch the data again from SEC and update the cache.
# If the request fails, it raises a RuntimeError with the HTTP status code and the first 200 characters of the response body.

def load_ticker_map():
    if os.path.exists(TICKER_CACHE_PATH):
        if time.time() - os.path.getmtime(TICKER_CACHE_PATH) < TICKER_CACHE_MAX_AGE:
            with open(TICKER_CACHE_PATH) as f:
                return json.load(f)
    resp = sec_get("https://www.sec.gov/files/company_tickers.json")
    if resp.status_code != 200:
        raise RuntimeError(f"Ticker registry request failed: HTTP {resp.status_code} — "
                            f"body starts with: {resp.text[:200]!r}")
    data = resp.json()
    with open(TICKER_CACHE_PATH, 'w') as f:
        json.dump(data, f)
    return data

# Resolves a ticker symbol to its corresponding CIK and company name using the ticker map.
# If the ticker is not found, it returns (None, None).
def resolve_cik(ticker, ticker_map):
    for item in ticker_map.values():
        if item['ticker'] == ticker:
            return str(item['cik_str']).zfill(10), item['title']
    return None, None

# Fetches company facts from the SEC's XBRL API for a given CIK.
# If the request fails, it raises a RuntimeError with the HTTP status code and the first 200 characters of the response body.
# The function returns the JSON response as a Python dictionary.
def fetch_company_facts(cik):
    # One bulk request per ticker instead of ~30+ per-concept requests —
    # this is both far faster and avoids tripping SEC's rate limiting.
    url = f"https://data.sec.gov/api/xbrl/companyfacts/CIK{cik}.json"
    resp = sec_get(url)
    if resp.status_code != 200:
        raise RuntimeError(f"companyfacts request failed: HTTP {resp.status_code} — "
                            f"body starts with: {resp.text[:200]!r}")
    return resp.json()


# =====================================================================
# CONCEPT TAXONOMIES (fallback tags per line item)
# =====================================================================
# The SEC's XBRL filings are inconsistent in which tags they use for a given line item, so we maintain a list of fallback tags for each concept. The first tag that is found in the filing is used.
# In future variations, may be best to use a more sophisticated mapping or even a machine learning model to identify the correct tags for each line item.
# May also be best to run through full list and record all tage found for each concept, to ensure that we are capturing all relevant data and not missing any important information.
CONCEPT_TAXONOMIES = {
    "Total Revenue": ["RevenueFromContractWithCustomerExcludingAssessedTax", "Revenues", "SalesRevenueNet", "SalesRevenueGoodsNet"],
    "Cost of Revenue": ["CostOfGoodsAndServicesSold", "CostOfRevenue", "CostOfGoodsSold"],
    "SG&A": ["SellingGeneralAndAdministrativeExpense"],
    "R&D": ["ResearchAndDevelopmentExpense"],
    "Operating Income": ["OperatingIncomeLoss"],
    "Interest Expense": ["InterestExpense", "InterestExpenseDebt"],
    "Pretax Income": ["IncomeLossFromContinuingOperationsBeforeIncomeTaxesExtraordinaryItemsNoncontrollingInterest",
                       "IncomeLossFromContinuingOperationsBeforeIncomeTaxesMinorityInterestAndIncomeLossFromEquityMethodInvestments"],
    "Income Tax Expense": ["IncomeTaxExpenseBenefit"],
    "Net Income": ["NetIncomeLoss"],
    "Diluted EPS": ["EarningsPerShareDiluted"],
    "Diluted Shares Outstanding (mm)": ["WeightedAverageNumberOfDilutedSharesOutstanding"],

    "Cash & Equivalents": ["CashAndCashEquivalentsAtCarryingValue", "CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalents"],
    "Accounts Receivable": ["AccountsReceivableNetCurrent"],
    "Inventory": ["InventoryNet"],
    "Total Current Assets": ["AssetsCurrent"],
    "PP&E, Net": ["PropertyPlantAndEquipmentNet"],
    "Goodwill": ["Goodwill"],
    "Total Assets": ["Assets"],
    "Accounts Payable": ["AccountsPayableCurrent"],
    "Total Current Liabilities": ["LiabilitiesCurrent"],
    "Long-Term Debt": ["LongTermDebtNoncurrent"],
    "Total Liabilities": ["Liabilities"],
    "Retained Earnings": ["RetainedEarningsAccumulatedDeficit"],
    "Total Stockholders Equity": ["StockholdersEquity"],

    "D&A": ["DepreciationDepletionAndAmortization", "DepreciationAmortizationAndAccretionNet", "DepreciationAndAmortization"],
    "Stock-Based Compensation": ["ShareBasedCompensation"],
    "Operating Cash Flow": ["NetCashProvidedByUsedInOperatingActivities"],
    "CapEx": ["PaymentsToAcquirePropertyPlantAndEquipment"],
    "Investing Cash Flow": ["NetCashProvidedByUsedInInvestingActivities"],
    "Dividends Paid": ["PaymentsOfDividends", "PaymentsOfDividendsCommonStock"],
    "Financing Cash Flow": ["NetCashProvidedByUsedInFinancingActivities"],
}

PER_SHARE_LABELS = {"Diluted EPS"}
SHARE_COUNT_LABELS = {"Diluted Shares Outstanding (mm)"}
MIN_ANNUAL_PERIOD_DAYS = 300  # excludes quarterly/stub periods that leak into a 10-K's units array

# Creates function that extracts annual series data from the SEC's XBRL filing JSON for a given set of tags and a timeline of years. It buckets each fact by the year the period actually covers (from 'end'), not by the filing's 'fy' tag. It also deduplicates by keeping the most recently filed figure for a given year, and restricts to 10-K/10-K-A annual periods.

def extract_annual_series(facts_json, tags, timeline):
    """
    Bucket each fact by the YEAR THE PERIOD ACTUALLY COVERS (from 'end'),
    not by the filing's 'fy' tag — a single 10-K reports several years of
    comparative figures, all stamped with the same 'fy'. Also dedupes by
    keeping the most-recently-filed figure for a given year (handles
    restatements/amendments), and restricts to 10-K/10-K-A annual periods.
    """
    gaap = facts_json.get('facts', {}).get('us-gaap', {})
    best = {}  # year -> (filed_date, value)

# Normalises the units to USD or USD/shares, and if neither is present, uses the first available unit. This is important because some filings may report in different currencies or units, and we want to ensure consistency in our extracted data.
    for tag in tags:
        concept = gaap.get(tag)
        if not concept:
            continue
        units = concept.get('units', {})
        unit_key = 'USD' if 'USD' in units else ('USD/shares' if 'USD/shares' in units else next(iter(units), None))
        if not unit_key:
            continue

# Iterates through the records for the selected unit, filtering for 10-K and 10-K/A forms, and extracting the value and end date. It checks if the end date's year is in the timeline, and if the period is at least 300 days (to exclude quarterly/stub periods). It keeps track of the most recently filed value for each year.
        for rec in units[unit_key]:
            if rec.get('form') not in ('10-K', '10-K/A'):
                continue
            val = rec.get('val')  # NOTE: SEC's field is 'val', not 'value'
            if val is None:
                continue
            end = rec.get('end')
            if not end:
                continue
            year = end[:4]
            if year not in timeline:
                continue

            start = rec.get('start')
            if start:
                days = (datetime.date.fromisoformat(end) - datetime.date.fromisoformat(start)).days
                if days < MIN_ANNUAL_PERIOD_DAYS:
                    continue  # quarterly/stub period, not the full-year figure

            filed = rec.get('filed', '')
            if year not in best or filed >= best[year][0]:
                best[year] = (filed, val)

    return {yr: v for yr, (f, v) in best.items()}

# Function that pulls company data for a given ticker and lookback period. It resolves the CIK and company name, fetches the company facts, and extracts the annual series data for each concept in the CONCEPT_TAXONOMIES. It scales the values appropriately based on whether they are per-share, share count, or other financial metrics. It returns the CIK, company name, timeline, extracted data, and coverage information.
def pull_company_data(ticker, lookback):
    ticker_map = load_ticker_map()
    cik, company_name = resolve_cik(ticker, ticker_map)
    if not cik:
        raise ValueError(f"Ticker '{ticker}' not found in SEC's registry.")

    facts = fetch_company_facts(cik)

    current_year = datetime.datetime.now().year
    timeline = [str(y) for y in range(current_year - lookback, current_year)]

    data = {}
    coverage = {}
    for label, tags in CONCEPT_TAXONOMIES.items():
        series = extract_annual_series(facts, tags, timeline)
        if label in PER_SHARE_LABELS:
            scaled = {yr: round(v, 2) for yr, v in series.items()}
        elif label in SHARE_COUNT_LABELS:
            scaled = {yr: round(v / 1_000_000, 1) for yr, v in series.items()}
        else:
            scaled = {yr: round(v / 1_000_000, 2) for yr, v in series.items()}
        data[label] = scaled
        coverage[label] = len(scaled)

    return cik, company_name, timeline, data, coverage


# =====================================================================
# EXCEL MODEL BUILDER
# =====================================================================
# Excel formatting constants in line with inductry standards for financial models. These include font styles, colors, number formats, and fill patterns for different sections of the model.
BLUE = Font(name='Arial', size=10, color='0000FF')
BLACK = Font(name='Arial', size=10, color='000000')
BLACK_BOLD = Font(name='Arial', size=10, color='000000', bold=True)
WHITE_BOLD = Font(name='Arial', size=11, color='FFFFFF', bold=True)
TITLE_FONT = Font(name='Arial', size=14, bold=True)
SECTION_FILL = PatternFill('solid', start_color='1F4E78', end_color='1F4E78')
SUBTOTAL_TOP_BORDER = Border(top=Side(style='thin'))
CHECK_FILL = PatternFill('solid', start_color='E2EFDA', end_color='E2EFDA')

# Number formats for different types of financial data, including currency, percentages, ratios, and shares. These formats are used to ensure consistency and readability in the Excel model.
NUM_FMT = '$#,##0;($#,##0);"-"'
EPS_FMT = '$#,##0.00;($#,##0.00);"-"'
SHARES_FMT = '#,##0.0;(#,##0.0);"-"'
PCT_FMT = '0.0%;(0.0%);"-"'
RATIO_FMT = '0.00x;(0.00x);"-"'

# Normalises terms into a standard set of line items for the income statement, balance sheet, cash flow statement, and key ratios. Each line item is associated with a type (raw data, formula, etc.) and a number format for display in the Excel model.   
INCOME_STATEMENT = [
    ("Total Revenue", "raw", NUM_FMT),
    ("Cost of Revenue", "raw", NUM_FMT),
    ("Gross Profit", "formula_gp", NUM_FMT),
    ("Gross Margin %", "formula_gm", PCT_FMT),
    ("SG&A", "raw", NUM_FMT),
    ("R&D", "raw", NUM_FMT),
    ("Operating Income", "raw", NUM_FMT),
    ("Operating Margin %", "formula_om", PCT_FMT),
    ("Interest Expense", "raw", NUM_FMT),
    ("Pretax Income", "raw", NUM_FMT),
    ("Income Tax Expense", "raw", NUM_FMT),
    ("Net Income", "raw", NUM_FMT),
    ("Net Margin %", "formula_nm", PCT_FMT),
    ("Revenue Growth % YoY", "formula_growth", PCT_FMT),
    ("Diluted EPS", "raw_eps", EPS_FMT),
    ("Diluted Shares Outstanding (mm)", "raw_shares", SHARES_FMT),
]
BALANCE_SHEET = [
    ("Cash & Equivalents", "raw", NUM_FMT),
    ("Accounts Receivable", "raw", NUM_FMT),
    ("Inventory", "raw", NUM_FMT),
    ("Total Current Assets", "raw", NUM_FMT),
    ("PP&E, Net", "raw", NUM_FMT),
    ("Goodwill", "raw", NUM_FMT),
    ("Total Assets", "raw", NUM_FMT),
    ("Accounts Payable", "raw", NUM_FMT),
    ("Total Current Liabilities", "raw", NUM_FMT),
    ("Long-Term Debt", "raw", NUM_FMT),
    ("Total Liabilities", "raw", NUM_FMT),
    ("Retained Earnings", "raw", NUM_FMT),
    ("Total Stockholders Equity", "raw", NUM_FMT),
    ("Total Liabilities & Equity", "formula_le", NUM_FMT),
    ("Balance Check (should be 0)", "formula_check", NUM_FMT),
]
CASH_FLOW = [
    ("Net Income", "link_ni", NUM_FMT),
    ("D&A", "raw", NUM_FMT),
    ("Stock-Based Compensation", "raw", NUM_FMT),
    ("Operating Cash Flow", "raw", NUM_FMT),
    ("CapEx", "raw", NUM_FMT),
    ("Investing Cash Flow", "raw", NUM_FMT),
    ("Dividends Paid", "raw", NUM_FMT),
    ("Financing Cash Flow", "raw", NUM_FMT),
    ("Net Change in Cash", "formula_netchg", NUM_FMT),
]
RATIOS = [
    ("Current Ratio", "formula_curr", RATIO_FMT),
    ("Debt / Equity", "formula_de", RATIO_FMT),
    ("Return on Equity (ROE)", "formula_roe", PCT_FMT),
    ("Return on Assets (ROA)", "formula_roa", PCT_FMT),
]

# Function that writes a section header in the Excel worksheet. It takes the worksheet, the current row, the title of the section, and the number of years to span. It formats the header with a bold white font on a colored background and fills the cells for the specified number of years with the same background color. It returns the next row number after the header.
def _write_section_header(ws, row, title, n_years):
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = WHITE_BOLD
    cell.fill = SECTION_FILL
    for c in range(2, 2 + n_years):
        ws.cell(row=row, column=c).fill = SECTION_FILL
    return row + 1

# Function that builds an Excel model for a given company based on its financial data. It takes the ticker symbol, company name, CIK, extracted data, timeline of years, and the output path for the Excel file. It creates a workbook, formats the worksheet, writes the income statement, balance sheet, cash flow statement, and key ratios using the provided data and formulas. Finally, it saves the workbook to the specified output path and returns the path.
def build_excel_model(ticker, company_name, cik, data, timeline, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Model"

    n_years = len(timeline)
    first_col = 2

    ws.cell(row=1, column=1, value=f"{company_name} ({ticker}) — Three-Statement Model").font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"CIK {cik}  |  $ in millions except per-share data  |  Source: SEC EDGAR")

    row = 4
    for i, yr in enumerate(timeline):
        c = ws.cell(row=row, column=first_col + i, value=yr)
        c.font = BLACK_BOLD
        c.alignment = Alignment(horizontal='center')
        c.number_format = '@'
    header_row = row
    row += 1

    line_row = {}

    def write_raw_row(label, fmt):
        nonlocal row
        ws.cell(row=row, column=1, value=label).font = BLACK
        for i, yr in enumerate(timeline):
            val = data.get(label, {}).get(yr)
            cell = ws.cell(row=row, column=first_col + i)
            if val is not None:
                cell.value = val
                cell.font = BLUE
            cell.number_format = fmt
        line_row[label] = row
        row += 1

# IMPORTANT function that writes a formula into EXCEL to preserve functionality of spreadsheet for users
    def write_formula_row(label, fmt, formula_fn, bold=False, top_border=False, fill_check=False):
        nonlocal row
        lbl_cell = ws.cell(row=row, column=1, value=label)
        lbl_cell.font = BLACK_BOLD if bold else BLACK
        for i in range(n_years):
            col = first_col + i
            cell = ws.cell(row=row, column=col)
            f = formula_fn(col, i)
            if f is not None:
                cell.value = f
            cell.font = BLACK_BOLD if bold else BLACK
            cell.number_format = fmt
            if top_border:
                cell.border = SUBTOTAL_TOP_BORDER
            if fill_check:
                cell.fill = CHECK_FILL
        if top_border:
            lbl_cell.border = SUBTOTAL_TOP_BORDER
        line_row[label] = row
        row += 1

    def L(label, col):
        return f"{get_column_letter(col)}{line_row[label]}"

    row = _write_section_header(ws, row, "INCOME STATEMENT", n_years)
    for label, kind, fmt in INCOME_STATEMENT:
        if kind in ("raw", "raw_eps", "raw_shares"):
            write_raw_row(label, fmt)
        elif kind == "formula_gp":
            write_formula_row(label, fmt, lambda c, i: f"={L('Total Revenue', c)}-{L('Cost of Revenue', c)}")
        elif kind == "formula_gm":
            write_formula_row(label, fmt, lambda c, i: f"=IF({L('Total Revenue', c)}=0,\"-\",{L('Gross Profit', c)}/{L('Total Revenue', c)})")
        elif kind == "formula_om":
            write_formula_row(label, fmt, lambda c, i: f"=IF({L('Total Revenue', c)}=0,\"-\",{L('Operating Income', c)}/{L('Total Revenue', c)})")
        elif kind == "formula_nm":
            write_formula_row(label, fmt, lambda c, i: f"=IF({L('Total Revenue', c)}=0,\"-\",{L('Net Income', c)}/{L('Total Revenue', c)})")
        elif kind == "formula_growth":
            def growth_f(c, i):
                if i == 0:
                    return None
                pc = c - 1
                return f"=IF({L('Total Revenue', pc)}=0,\"-\",({L('Total Revenue', c)}-{L('Total Revenue', pc)})/{L('Total Revenue', pc)})"
            write_formula_row(label, fmt, growth_f)

    row += 1
    row = _write_section_header(ws, row, "BALANCE SHEET", n_years)
    for label, kind, fmt in BALANCE_SHEET:
        if kind == "raw":
            write_raw_row(label, fmt)
        elif kind == "formula_le":
            write_formula_row(label, fmt, lambda c, i: f"={L('Total Liabilities', c)}+{L('Total Stockholders Equity', c)}")
        elif kind == "formula_check":
            write_formula_row(label, fmt, lambda c, i: f"=ROUND({L('Total Assets', c)}-{L('Total Liabilities & Equity', c)},0)", fill_check=True)

    row += 1
    row = _write_section_header(ws, row, "CASH FLOW STATEMENT", n_years)
    for label, kind, fmt in CASH_FLOW:
        if kind == "raw":
            write_raw_row(label, fmt)
        elif kind == "link_ni":
            write_formula_row(label, fmt, lambda c, i: f"={L('Net Income', c)}")
        elif kind == "formula_netchg":
            write_formula_row(label, fmt, lambda c, i: f"={L('Operating Cash Flow', c)}+{L('Investing Cash Flow', c)}+{L('Financing Cash Flow', c)}", bold=True, top_border=True)

    row += 1
    row = _write_section_header(ws, row, "KEY RATIOS", n_years)
    for label, kind, fmt in RATIOS:
        if kind == "formula_curr":
            write_formula_row(label, fmt, lambda c, i: f"=IF({L('Total Current Liabilities', c)}=0,\"-\",{L('Total Current Assets', c)}/{L('Total Current Liabilities', c)})")
        elif kind == "formula_de":
            write_formula_row(label, fmt, lambda c, i: f"=IF({L('Total Stockholders Equity', c)}=0,\"-\",{L('Long-Term Debt', c)}/{L('Total Stockholders Equity', c)})")
        elif kind == "formula_roe":
            write_formula_row(label, fmt, lambda c, i: f"=IF({L('Total Stockholders Equity', c)}=0,\"-\",{L('Net Income', c)}/{L('Total Stockholders Equity', c)})")
        elif kind == "formula_roa":
            write_formula_row(label, fmt, lambda c, i: f"=IF({L('Total Assets', c)}=0,\"-\",{L('Net Income', c)}/{L('Total Assets', c)})")

    ws.column_dimensions['A'].width = 32
    for i in range(n_years):
        ws.column_dimensions[get_column_letter(first_col + i)].width = 13
    ws.freeze_panes = ws.cell(row=header_row + 1, column=first_col)

    wb.save(output_path)
    return output_path

# Auto-open function in native spreadsheet workbook
def open_file_in_default_app(path):
    """Best-effort: open the saved workbook in the OS's default app (Excel, etc.)."""
    try:
        system = platform.system()
        if system == 'Windows':
            os.startfile(path)  # noqa: only exists on Windows
        elif system == 'Darwin':
            subprocess.run(['open', path], check=False)
        else:
            subprocess.run(['xdg-open', path], check=False)
        return True
    except Exception as e:
        print(f"   ⚠️ Couldn't auto-open the file ({e}). Open it manually from: {path}")
        return False


# =====================================================================
# UI LAYOUT
# =====================================================================
# UI elements - formatting for UI in jupyter notebook
ticker_box = widgets.Text(value='ILMN', placeholder='Ticker...', description='Ticker:', layout=widgets.Layout(width='30%'))
years_box = widgets.IntText(value=10, description='Years:', layout=widgets.Layout(width='20%'))
action_btn = widgets.Button(description='Build Model Matrix', button_style='success', icon='calculator', layout=widgets.Layout(width='30%'))
display_output = widgets.Output()

# Function that handles the button click event to pull the financial matrix for the specified ticker and lookback period. It clears the output, retrieves the ticker and lookback values from the UI elements, validates the ticker, and then calls the pull_company_data function to fetch the data. It displays the results, including coverage information, and builds the Excel model using the build_excel_model function. Finally, it saves the model and provides a link to open it.
def pull_financial_matrix(change_event):
    with display_output:
        clear_output()
        ticker = ticker_box.value.strip().upper()
        lookback = years_box.value

        if not ticker:
            print("⚠️ Please provide a valid stock ticker symbol.")
            return

        print(f"📡 Resolving CIK + pulling SEC company facts for: {ticker}...")
        try:
            cik, company_name, timeline, data, coverage = pull_company_data(ticker, lookback)
        except Exception as e:
            print(f"❌ {e}")
            return

        print(f"   🎯 Match: {company_name} | CIK: {cik}")
        print("-" * 80)
        for label, n in coverage.items():
            mark = "✅" if n > 0 else "⚠️"
            print(f"   {mark} [{n:>2}/{len(timeline)} yrs] {label}")
        print("-" * 80)

        out_path = os.path.abspath(f"{ticker}_three_statement_model.xlsx")
        build_excel_model(ticker, company_name, cik, data, timeline, out_path)
        print(f"\n💎 Model saved: {out_path}")
        display(FileLink(out_path))
        open_file_in_default_app(out_path)


action_btn.on_click(pull_financial_matrix)
ui_panel = widgets.HBox([ticker_box, years_box, action_btn])
print("📈 Multi-Concept SEC Financial Modeling System")
display(ui_panel, display_output)
# All neatly encased within a button click event
